"""
Export clinical trials data from SQLite to Excel.
Sheets: All Trials, Summary by Company, Summary by Phase, Summary by Country + Charts.
"""
import sqlite3
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

DB_PATH = "trials.db"
OUTPUT_PATH = "clinical_trials_data.xlsx"

BRAND_BLUE = "1a1a2e"
ACCENT_BLUE = "0093D0"
LIGHT_BLUE = "D6EAF8"
LIGHT_GRAY = "F2F3F4"


def style_header_row(ws, row=1, fill_color=BRAND_BLUE, font_color="FFFFFF"):
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(bold=True, color=font_color, size=11)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_rows(ws, start_row=2):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(bottom=thin, right=thin)
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        fill_color = LIGHT_BLUE if i % 2 == 0 else "FFFFFF"
        fill = PatternFill(fill_type="solid", fgColor=fill_color)
        for cell in row:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_fit_columns(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


conn = sqlite3.connect(DB_PATH)

FOCUS_COUNTRIES = ('Israel','Belgium','Switzerland','Austria','Sweden','Denmark','Norway','Ireland')
COUNTRY_IN = ",".join(f"'{c}'" for c in FOCUS_COUNTRIES)

# --- Sheet 1: All Trials ---
df_trials = pd.read_sql_query(f"""
    SELECT
        t.nct_id                                    AS "NCT ID",
        t.company                                   AS "Company",
        t.sponsor                                   AS "Sponsor (ClinicalTrials.gov)",
        REPLACE(t.phase, 'PHASE', 'Phase ')         AS "Phase",
        t.start_year                                AS "Start Year",
        t.start_month                               AS "Start Month",
        GROUP_CONCAT(tc.country || ' (' || COALESCE(tc.site_count,1) || ')', ', ')
                                                    AS "Countries (sites per country)"
    FROM trials t
    LEFT JOIN trial_countries tc ON t.nct_id = tc.nct_id
    GROUP BY t.nct_id
    ORDER BY t.start_year DESC, t.start_month DESC
""", conn)

# --- Sheet 2: Summary by Company & Year ---
df_company_year = pd.read_sql_query(f"""
    SELECT
        t.company                                   AS "Company",
        t.start_year                                AS "Year",
        REPLACE(t.phase, 'PHASE', 'Phase ')         AS "Phase",
        COUNT(DISTINCT t.nct_id)                    AS "Number of Trials",
        CAST(SUM(COALESCE(tc.site_count, 1)) AS INTEGER)
                                                    AS "Total Research Sites"
    FROM trials t
    JOIN trial_countries tc ON t.nct_id = tc.nct_id
    WHERE tc.country IN ({COUNTRY_IN})
    GROUP BY t.company, t.start_year, t.phase
    ORDER BY t.company, t.start_year, t.phase
""", conn)

# --- Sheet 3: Summary by Company (pivot — trial counts) ---
df_pivot_raw = pd.read_sql_query("""
    SELECT company, start_year, COUNT(*) AS cnt
    FROM trials
    GROUP BY company, start_year
    ORDER BY company, start_year
""", conn)
df_pivot = df_pivot_raw.pivot(index="company", columns="start_year", values="cnt").fillna(0).astype(int)
df_pivot["Total"] = df_pivot.sum(axis=1)
df_pivot = df_pivot.sort_values("Total", ascending=False).reset_index()
df_pivot.columns.name = None
df_pivot = df_pivot.rename(columns={"company": "Company"})

# --- Sheet 4: Summary by Country ---
df_country = pd.read_sql_query(f"""
    SELECT
        tc.country                                          AS "Country",
        t.company                                          AS "Company",
        REPLACE(t.phase, 'PHASE', 'Phase ')                AS "Phase",
        COUNT(DISTINCT t.nct_id)                           AS "Number of Trials",
        CAST(SUM(COALESCE(tc.site_count, 1)) AS INTEGER)   AS "Total Research Sites"
    FROM trial_countries tc
    JOIN trials t ON tc.nct_id = t.nct_id
    WHERE tc.country IN ({COUNTRY_IN})
    GROUP BY tc.country, t.company, t.phase
    ORDER BY tc.country, t.company, t.phase
""", conn)

# --- Sheet 5: Country Pivot (trial counts) ---
df_country_pivot_raw = pd.read_sql_query(f"""
    SELECT tc.country AS "Country", t.company AS "Company", COUNT(DISTINCT t.nct_id) AS cnt
    FROM trial_countries tc
    JOIN trials t ON tc.nct_id = t.nct_id
    WHERE tc.country IN ({COUNTRY_IN})
    GROUP BY tc.country, t.company
""", conn)
df_country_pivot = df_country_pivot_raw.pivot(index="Country", columns="Company", values="cnt").fillna(0).astype(int)
df_country_pivot["Total Trials"] = df_country_pivot.sum(axis=1)
df_country_pivot = df_country_pivot.sort_values("Total Trials", ascending=False).reset_index()
df_country_pivot.columns.name = None

# --- Sheet 6: Country Pivot (research sites) ---
df_sites_pivot_raw = pd.read_sql_query(f"""
    SELECT
        tc.country                                          AS "Country",
        t.company                                          AS "Company",
        CAST(SUM(COALESCE(tc.site_count, 1)) AS INTEGER)   AS sites
    FROM trial_countries tc
    JOIN trials t ON tc.nct_id = t.nct_id
    WHERE tc.country IN ({COUNTRY_IN})
    GROUP BY tc.country, t.company
""", conn)
df_sites_pivot = df_sites_pivot_raw.pivot(index="Country", columns="Company", values="sites").fillna(0).astype(int)
df_sites_pivot["Total Sites"] = df_sites_pivot.sum(axis=1)
df_sites_pivot = df_sites_pivot.sort_values("Total Sites", ascending=False).reset_index()
df_sites_pivot.columns.name = None

CHART_COUNTRIES = ('Israel', 'Belgium', 'Switzerland', 'Austria', 'Sweden', 'Denmark', 'Norway', 'Ireland')
CHART_YEARS     = list(range(2015, 2026))

# Population in millions — World Bank WDI (2024-2025: estimates)
POPULATION = {
    "Israel":      [8.38,  8.54,  8.71,  8.88,  9.05,  9.22,  9.45,  9.66,  9.84,  10.01, 10.20],
    "Belgium":     [11.24, 11.31, 11.35, 11.43, 11.51, 11.59, 11.59, 11.66, 11.74, 11.87, 11.95],
    "Switzerland": [8.33,  8.40,  8.48,  8.55,  8.60,  8.67,  8.74,  8.82,  8.96,  9.10,  9.18],
    "Austria":     [8.66,  8.74,  8.80,  8.85,  8.90,  8.93,  9.04,  9.10,  9.10,  9.15,  9.21],
    "Sweden":      [9.80,  9.92,  10.07, 10.18, 10.29, 10.33, 10.42, 10.52, 10.55, 10.65, 10.75],
    "Denmark":     [5.69,  5.73,  5.75,  5.80,  5.81,  5.82,  5.84,  5.88,  5.96,  6.03,  6.10],
    "Norway":      [5.19,  5.23,  5.27,  5.30,  5.33,  5.37,  5.41,  5.43,  5.52,  5.59,  5.67],
    "Ireland":     [4.63,  4.71,  4.78,  4.86,  4.94,  5.00,  5.13,  5.24,  5.31,  5.41,  5.51],
}

COUNTRY_COLORS_HEX = {
    "Israel":      "0038B8",
    "Belgium":     "E63946",
    "Switzerland": "2A9D8F",
    "Austria":     "E9C46A",
    "Sweden":      "264653",
    "Denmark":     "F4A261",
    "Norway":      "457B9D",
    "Ireland":     "6A994E",
}

CHART_SPECS = [
    # (data_sheet,     chart_sheet,              chart_title,                              y_label,           phase,    metric,  per_capita)
    ("D-Ph1 Trials",  "Chart - Ph1 Trials",      "Phase 1 — Trials per Year",             "Number of Trials", "PHASE1", "trials", False),
    ("D-Ph2 Trials",  "Chart - Ph2 Trials",      "Phase 2 — Trials per Year",             "Number of Trials", "PHASE2", "trials", False),
    ("D-Ph3 Trials",  "Chart - Ph3 Trials",      "Phase 3 — Trials per Year",             "Number of Trials", "PHASE3", "trials", False),
    ("D-All Trials",  "Chart - All Trials",       "All Phases — Trials per Year",          "Number of Trials", None,     "trials", False),
    ("D-Ph1 Sites",   "Chart - Ph1 Sites",        "Phase 1 — Research Sites per Year",    "Research Sites",   "PHASE1", "sites",  False),
    ("D-Ph2 Sites",   "Chart - Ph2 Sites",        "Phase 2 — Research Sites per Year",    "Research Sites",   "PHASE2", "sites",  False),
    ("D-Ph3 Sites",   "Chart - Ph3 Sites",        "Phase 3 — Research Sites per Year",    "Research Sites",   "PHASE3", "sites",  False),
    ("D-All Sites",   "Chart - All Sites",         "All Phases — Research Sites per Year", "Research Sites",   None,     "sites",  False),
    # Per-capita versions (per million population)
    ("D-Ph1 Tr pM",   "Chart - Ph1 Trials pM",   "Phase 1 — Trials per Million Pop.",    "Trials per M Pop.", "PHASE1", "trials", True),
    ("D-Ph2 Tr pM",   "Chart - Ph2 Trials pM",   "Phase 2 — Trials per Million Pop.",    "Trials per M Pop.", "PHASE2", "trials", True),
    ("D-Ph3 Tr pM",   "Chart - Ph3 Trials pM",   "Phase 3 — Trials per Million Pop.",    "Trials per M Pop.", "PHASE3", "trials", True),
    ("D-All Tr pM",   "Chart - All Trials pM",    "All Phases — Trials per Million Pop.", "Trials per M Pop.", None,     "trials", True),
    ("D-Ph1 Si pM",   "Chart - Ph1 Sites pM",     "Phase 1 — Sites per Million Pop.",     "Sites per M Pop.",  "PHASE1", "sites",  True),
    ("D-Ph2 Si pM",   "Chart - Ph2 Sites pM",     "Phase 2 — Sites per Million Pop.",     "Sites per M Pop.",  "PHASE2", "sites",  True),
    ("D-Ph3 Si pM",   "Chart - Ph3 Sites pM",     "Phase 3 — Sites per Million Pop.",     "Sites per M Pop.",  "PHASE3", "sites",  True),
    ("D-All Si pM",   "Chart - All Sites pM",      "All Phases — Sites per Million Pop.", "Sites per M Pop.",  None,     "sites",  True),
]


def get_chart_pivot(conn, phase, metric, per_capita=False):
    """Return DataFrame with Year as first column and one column per country."""
    country_in_str = ",".join(f"'{c}'" for c in CHART_COUNTRIES)
    phase_clause   = f"AND t.phase = '{phase}'" if phase else \
                     "AND t.phase IN ('PHASE1','PHASE2','PHASE3')"
    value_expr     = ("COUNT(DISTINCT t.nct_id)"
                      if metric == "trials"
                      else "CAST(SUM(COALESCE(tc.site_count, 1)) AS INTEGER)")

    df = pd.read_sql_query(f"""
        SELECT t.start_year AS year, tc.country, {value_expr} AS value
        FROM trials t
        JOIN trial_countries tc ON t.nct_id = tc.nct_id
        WHERE t.start_year BETWEEN 2015 AND 2025
          {phase_clause}
          AND tc.country IN ({country_in_str})
        GROUP BY t.start_year, tc.country
    """, conn)

    pivot = (df.pivot(index="year", columns="country", values="value")
               .reindex(index=CHART_YEARS, columns=list(CHART_COUNTRIES))
               .fillna(0).astype(int))

    if per_capita:
        pivot = pivot.astype(float)
        for i, yr in enumerate(CHART_YEARS):
            for country in CHART_COUNTRIES:
                pop = POPULATION[country][i]
                pivot.at[yr, country] = round(float(pivot.at[yr, country]) / pop, 3) if pop else 0.0

    pivot.index.name = "Year"
    return pivot.reset_index()


def write_chart_data_sheet(wb, sheet_name, df_pivot):
    """Write pivot data to a sheet; return the worksheet."""
    ws = wb.create_sheet(sheet_name)
    # Header
    header_fill = PatternFill(fill_type="solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, col_name in enumerate(df_pivot.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Data rows
    for r_idx, row in enumerate(df_pivot.itertuples(index=False), start=2):
        alt_fill = PatternFill(fill_type="solid", fgColor="D6EAF8" if r_idx % 2 == 0 else "FFFFFF")
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="center")
    # Column widths
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 14
    ws.row_dimensions[1].height = 28
    return ws


def add_line_chart(wb, data_sheet_name, chart_sheet_name, chart_title, y_label):
    """Create a line chart sheet referencing the data sheet."""
    ws_data = wb[data_sheet_name]
    n_rows  = len(CHART_YEARS) + 1   # +1 for header

    chart              = LineChart()
    chart.title        = chart_title
    chart.style        = 10
    chart.y_axis.title = y_label
    chart.x_axis.title = "Year"
    chart.height       = 18
    chart.width        = 32

    # Add all country columns at once (cols 2..9)
    n_cols = len(CHART_COUNTRIES)
    data_ref = Reference(ws_data, min_col=2, max_col=1 + n_cols, min_row=1, max_row=n_rows)
    chart.add_data(data_ref, titles_from_data=True)

    # X-axis categories
    cats = Reference(ws_data, min_col=1, max_col=1, min_row=2, max_row=n_rows)
    chart.set_categories(cats)

    # Color each series to match dashboard
    for series, country in zip(chart.series, CHART_COUNTRIES):
        hex_color = COUNTRY_COLORS_HEX[country]
        series.graphicalProperties.line.solidFill    = hex_color
        series.graphicalProperties.line.width        = 20000   # ~1.6pt
        series.marker.symbol                          = "circle"
        series.marker.size                            = 6
        series.marker.graphicalProperties.solidFill  = hex_color
        series.marker.graphicalProperties.line.solidFill = hex_color
        series.smooth = False

    chart.legend.position = "r"

    cs = wb.create_chartsheet(chart_sheet_name)
    cs.add_chart(chart)


conn.close()

# --- Write to Excel ---
with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
    data_sheets = [
        (df_trials,          "All Trials"),
        (df_company_year,    "By Company & Year"),
        (df_pivot,           "Company Pivot (trials)"),
        (df_country,         "By Country"),
        (df_country_pivot,   "Country Pivot (trials)"),
        (df_sites_pivot,     "Country Pivot (sites)"),
    ]

    for df, sheet_name in data_sheets:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
        ws = writer.sheets[sheet_name]
        style_header_row(ws)
        style_data_rows(ws)
        auto_fit_columns(ws)
        freeze_and_filter(ws)
        ws.row_dimensions[1].height = 30

    wb = writer.book
    conn2 = sqlite3.connect(DB_PATH)

    # Population sheet
    ws_pop = wb.create_sheet("Population (millions)")
    pop_header_fill = PatternFill(fill_type="solid", fgColor="1a1a2e")
    pop_header_font = Font(bold=True, color="FFFFFF", size=11)
    header_cells = ["Country"] + [str(y) for y in CHART_YEARS]
    for c_idx, h in enumerate(header_cells, 1):
        cell = ws_pop.cell(row=1, column=c_idx, value=h)
        cell.fill = pop_header_fill
        cell.font = pop_header_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, country in enumerate(CHART_COUNTRIES, 2):
        ws_pop.cell(row=r_idx, column=1, value=country).font = Font(bold=True)
        for c_idx, pop_val in enumerate(POPULATION[country], 2):
            ws_pop.cell(row=r_idx, column=c_idx, value=pop_val)
    ws_pop.column_dimensions["A"].width = 16
    for col in range(2, len(CHART_YEARS) + 2):
        ws_pop.column_dimensions[get_column_letter(col)].width = 9
    ws_pop.row_dimensions[1].height = 25
    ws_pop.append([])
    note_row = len(CHART_COUNTRIES) + 3
    ws_pop.cell(row=note_row, column=1,
                value="Source: World Bank World Development Indicators. 2024-2025: estimates.").font = Font(italic=True, color="888888")

    for data_sheet, chart_sheet, title, ylabel, phase, metric, per_capita in CHART_SPECS:
        df_pivot_chart = get_chart_pivot(conn2, phase, metric, per_capita)
        write_chart_data_sheet(wb, data_sheet, df_pivot_chart)
        add_line_chart(wb, data_sheet, chart_sheet, title, ylabel)
        print(f"  Added: {chart_sheet}")

    conn2.close()

print(f"\nExcel file saved: {OUTPUT_PATH}")
for _, name in data_sheets:
    print(f"  Sheet: {name}")
